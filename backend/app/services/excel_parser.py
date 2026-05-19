"""Streaming Excel parser using openpyxl read-only mode.

Reads row-by-row to keep memory constant for 100k+ row files.
Applies cleaning rules from the requirements doc:
- Percent strings (e.g. "8.97%") → decimal 0.0897
- Tab characters and whitespace stripped
- "其它" → "其他" normalized
- None/empty → 0 for numeric fields
- Date formats: YYYY-MM-DD, YYYY/MM/DD, YYYY.MM.DD
"""
from __future__ import annotations

import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Iterator

from openpyxl import load_workbook


COLUMN_MAP: dict[str, str] = {
    "店铺": "shop_name",
    "店铺编码": "shop_code",
    "日期": "date",
    "分类": "sku",
    "类目分类": "category",
    "负责人": "owner",
    "资料-实发数量(订单)": "order_qty",
    "售后-发货前退款占比": "pre_ship_refund_rate",
    "售后-发货后实退金额": "post_ship_refund_amt",
    "售后-发货后退平台补贴": "post_ship_platform_subsidy",
    "售后-发货后退款占比": "post_ship_refund_rate",
    "收入-实收金额": "actual_income",
    "收入-退款金额合计": "refund_total",
    "收入-退款占比": "refund_rate",
    "收入-收入总额": "income_total",
    "成本-耗材成本": "material_cost",
    "成本-耗材成本占比": "material_cost_rate",
    "成本-成本总额": "cost_total",
    "费用-赠品成本": "gift_cost",
    "费用-赠品成本占比": "gift_cost_rate",
    "费用-毛利率1": "gross_margin",
    "费用-营销费用": "marketing_cost",
    "费用-营销费用占比": "marketing_cost_rate",
    "费用-快递费用": "shipping_cost",
    "费用-快递费用占比": "shipping_cost_rate",
    "费用-平台费用": "platform_cost",
    "费用-平台费用占比": "platform_cost_rate",
    "利润-经营利润": "profit",
    "利润-经营利润率": "profit_rate",
}

PERCENT_FIELDS = {
    "pre_ship_refund_rate", "post_ship_refund_rate", "refund_rate",
    "material_cost_rate", "gift_cost_rate", "gross_margin",
    "marketing_cost_rate", "shipping_cost_rate", "platform_cost_rate",
    "profit_rate",
}

NUMERIC_FIELDS = PERCENT_FIELDS | {
    "post_ship_refund_amt", "post_ship_platform_subsidy",
    "actual_income", "refund_total", "income_total",
    "material_cost", "cost_total", "gift_cost",
    "marketing_cost", "shipping_cost", "platform_cost", "profit",
}

INT_FIELDS = {"order_qty"}
DIMENSION_FIELDS = {"shop_name", "shop_code", "sku", "category", "owner"}


def _clean_str(v) -> str:
    if v is None:
        return ""
    s = str(v).replace("\t", "").strip()
    return s


def _parse_percent(v) -> Decimal:
    """Convert percent string/number to a fractional decimal (8.97% → 0.0897)."""
    if v is None or v == "":
        return Decimal("0")
    if isinstance(v, (int, float, Decimal)):
        # If looks like already-fractional (between -1 and 1), use as-is
        d = Decimal(str(v))
        return d / Decimal("100") if abs(d) > Decimal("1") else d
    s = _clean_str(v).replace(",", "")
    if not s:
        return Decimal("0")
    has_pct = s.endswith("%")
    s = s.rstrip("%").strip()
    try:
        d = Decimal(s)
    except InvalidOperation:
        return Decimal("0")
    return d / Decimal("100") if has_pct else d


def _parse_number(v) -> Decimal:
    if v is None or v == "":
        return Decimal("0")
    if isinstance(v, (int, float, Decimal)):
        return Decimal(str(v))
    s = _clean_str(v).replace(",", "")
    if not s:
        return Decimal("0")
    try:
        return Decimal(s)
    except InvalidOperation:
        return Decimal("0")


def _parse_int(v) -> int:
    if v is None or v == "":
        return 0
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        return int(v)
    s = _clean_str(v).replace(",", "")
    if not s:
        return 0
    try:
        return int(float(s))
    except ValueError:
        return 0


_DATE_PATTERNS = [
    re.compile(r"^(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})$"),
]


def _parse_date(v) -> date | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    # Excel serial date: cells with General format containing a date come through
    # as float (days since 1899-12-30, ignoring Excel's 1900 leap-year bug).
    if isinstance(v, (int, float)) and 0 < v < 200000:
        try:
            from datetime import timedelta
            return (datetime(1899, 12, 30) + timedelta(days=float(v))).date()
        except (OverflowError, ValueError):
            return None
    s = _clean_str(v)
    for pat in _DATE_PATTERNS:
        m = pat.match(s)
        if m:
            y, mo, d = map(int, m.groups())
            try:
                return date(y, mo, d)
            except ValueError:
                return None
    return None


def _normalize_category(v: str) -> str:
    s = _clean_str(v)
    return s.replace("其它", "其他") if s else s


def parse_workbook(path: str | Path, stats: dict | None = None) -> Iterator[dict]:
    """Stream rows from the first sheet of an xlsx file.

    Yields dicts with cleaned & typed field values.

    If `stats` is provided, the parser records per-row skip reasons into it so
    the caller can tell the user how many rows were dropped and why. Keys:
      - scanned       : rows that had at least one non-empty cell
      - skipped_empty : rows where every cell was blank
      - skipped_no_keys : rows missing shop_code / date / sku
      - skipped_bad_date : rows whose date cell couldn't be parsed
    """
    if stats is not None:
        stats.setdefault("scanned", 0)
        stats.setdefault("skipped_empty", 0)
        stats.setdefault("skipped_no_keys", 0)
        stats.setdefault("skipped_bad_date", 0)

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            return
        # Map header positions → DB field names (loose: ignore unknown columns)
        col_idx: dict[str, int] = {}
        for i, h in enumerate(header):
            key = _clean_str(h)
            if key in COLUMN_MAP:
                col_idx[COLUMN_MAP[key]] = i

        for row in rows:
            if not row or all(v is None or v == "" for v in row):
                if stats is not None:
                    stats["skipped_empty"] += 1
                continue
            if stats is not None:
                stats["scanned"] += 1
            rec: dict = {}
            date_raw = None
            for field, idx in col_idx.items():
                raw = row[idx] if idx < len(row) else None
                if field == "date":
                    date_raw = raw
                    rec[field] = _parse_date(raw)
                elif field == "category":
                    rec[field] = _normalize_category(raw)
                elif field in DIMENSION_FIELDS:
                    rec[field] = _clean_str(raw)
                elif field in INT_FIELDS:
                    rec[field] = _parse_int(raw)
                elif field in PERCENT_FIELDS:
                    rec[field] = _parse_percent(raw)
                elif field in NUMERIC_FIELDS:
                    rec[field] = _parse_number(raw)
                else:
                    rec[field] = raw
            # Distinguish "date cell had content but unparseable" from "all keys missing".
            if rec.get("date") is None and date_raw not in (None, ""):
                if stats is not None:
                    stats["skipped_bad_date"] += 1
                continue
            if not rec.get("shop_code") or not rec.get("date") or not rec.get("sku"):
                if stats is not None:
                    stats["skipped_no_keys"] += 1
                continue
            yield rec
    finally:
        wb.close()
