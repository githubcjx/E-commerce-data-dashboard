"""Quick inspector for the user's real 数据源.xlsx — print sheets, headers, types, sample rows."""
from __future__ import annotations

import sys
from collections import Counter
from openpyxl import load_workbook


def main(path: str) -> None:
    wb = load_workbook(filename=path, read_only=True, data_only=True)
    print(f"file: {path}")
    print(f"sheets: {wb.sheetnames}")
    print()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"=== sheet: {sheet_name} (max_row={ws.max_row}, max_col={ws.max_column}) ===")
        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            print("  (empty)")
            continue

        print("HEADERS:")
        for i, h in enumerate(header):
            print(f"  [{i:2}] {h!r}")
        print()

        # First 3 data rows with types
        print("SAMPLE ROWS (first 3):")
        for ri, row in enumerate(rows):
            if ri >= 3:
                break
            for ci, (h, v) in enumerate(zip(header, row)):
                print(f"  row{ri+1}.[{ci:2}] {str(h)[:24]:24} = {v!r:60}  type={type(v).__name__}")
            print()

        # Sample stats: count unique shops, dates, categories — to validate aggregation
        rows2 = ws.iter_rows(values_only=True)
        next(rows2)  # skip header
        shops = Counter()
        owners_biz = Counter()
        categories = Counter()
        dates = Counter()
        skus = Counter()
        total = 0
        sample_pct_values = []
        for row in rows2:
            total += 1
            # naive: try to find columns by header name
            for i, h in enumerate(header):
                v = row[i] if i < len(row) else None
                if h is None: continue
                s = str(h).strip()
                if s == "店铺": shops[v] += 1
                elif s == "店铺编码": pass
                elif s == "日期": dates[str(v)] += 1
                elif s == "分类": skus[v] += 1
                elif s == "类目分类": categories[v] += 1
                elif s == "负责人": owners_biz[v] += 1
                elif "退款占比" in s and len(sample_pct_values) < 5:
                    sample_pct_values.append((s, v, type(v).__name__))
        print(f"TOTAL ROWS: {total}")
        print(f"unique shops:      {len(shops):3} → {list(shops)[:10]}")
        print(f"unique categories: {len(categories):3} → {list(categories)[:10]}")
        print(f"unique owners:     {len(owners_biz):3} → {list(owners_biz)[:10]}")
        print(f"unique dates:      {len(dates):3} → {sorted(dates)[:5]} ... {sorted(dates)[-3:]}")
        print(f"unique SKU(分类):  {len(skus):3}")
        if sample_pct_values:
            print("SAMPLE PERCENT VALUES:")
            for h, v, t in sample_pct_values:
                print(f"  {h}: {v!r} (type={t})")

        print()
    wb.close()


if __name__ == "__main__":
    main(sys.argv[1] if len(sys.argv) > 1 else r"C:\Users\Admin\Pictures\数据源.xlsx")
